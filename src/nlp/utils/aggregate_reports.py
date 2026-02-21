# src/nlp/utils/aggregate_reports_readable.py
"""
Aggregates radiology reports into a readable XLSX file.
- Preserves line breaks in report_text for natural reading
- Auto-adjusts column widths
- Adds basic flags (e.g., contains 'Impression:', 'Findings:')
- One row per report, wide report_text column

Usage:
    python src/nlp/utils/aggregate_reports.py
"""

from pathlib import Path
import pandas as pd
from datetime import datetime
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

REPORTS_DIR = Path("workspace/data/preprocessed_data/reports")
OUTPUT_DIR = Path("results")
OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

TIMESTAMP = datetime.now().strftime("%Y%m%d_%H%M")
OUTPUT_XLSX = OUTPUT_DIR / f"compiled_reports_readable_{TIMESTAMP}.xlsx"


def aggregate_readable():
    txt_files = sorted(REPORTS_DIR.glob("*.txt"))
    if not txt_files:
        print(f"No reports found in {REPORTS_DIR}")
        return

    print(f"Found {len(txt_files)} reports. Building readable table...")

    data = []
    for txt_file in txt_files:
        try:
            text = txt_file.read_text(encoding="utf-8", errors="replace").strip()
            data.append({
                "study_id": txt_file.stem,
                "report_text": text,
                "word_count": len(text.split()),
                "character_count": len(text),
                "has_findings": "Findings:" in text,
                "has_impression": any(word in text for word in ["Impression:", "Conclusion:", "CONCLUSION"])
            })
        except Exception as e:
            print(f"Error reading {txt_file.name}: {e}")

    df = pd.DataFrame(data)
    df = df.sort_values("study_id")

    # Save to XLSX with formatting
    wb = Workbook()
    ws = wb.active
    ws.title = "Reports"

    # Write headers
    headers = list(df.columns)
    ws.append(headers)

    # Write data rows
    for r_idx, row in enumerate(df.itertuples(index=False), start=2):
        for c_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            # Enable text wrapping for report_text column (column B = 2)
            if c_idx == 2:  # report_text
                cell.alignment = cell.alignment.copy(wrap_text=True)

    # Auto-adjust column widths
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        # Cap report_text column at reasonable width (adjust as needed)
        if column == "B":  # report_text
            adjusted_width = min(adjusted_width, 120)
        ws.column_dimensions[column].width = adjusted_width

    # Freeze header row
    ws.freeze_panes = "A2"

    wb.save(OUTPUT_XLSX)
    print(f"\nReadable XLSX created → {OUTPUT_XLSX}")
    print(f"Total reports: {len(df)}")
    print("Open in Excel/LibreOffice:")
    print("  • Column B (report_text) will show full reports with line breaks")
    print("  • Use Ctrl+F to search phrases across all reports")
    print("  • Filter on 'has_findings' or 'has_impression' to focus on key sections")

if __name__ == "__main__":
    aggregate_readable()